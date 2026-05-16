import streamlit as st
import pandas as pd
import requests
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import time
import random
from openpyxl.styles import PatternFill

# --- KONFIGURASI USER-AGENT ---
# Rotasi identitas browser agar tidak mudah diblokir Google
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) Gecko/20100101 Firefox/119.0"
]

# --- DAFTAR KODE BAHASA UMUM (639-1) ---
# Ditambahkan agar user mudah memilih, tapi tetap bisa manual input
COMMON_LANGUAGES = {
    "Inggris": "en",
    "Dari (Afghanistan)": "fa-AF",  # Kode spesifik untuk Dari
    "Arab": "ar",
    "Jepang": "ja",
    "Korea": "ko",
    "Mandarin": "zh-CN",
    "Prancis": "fr",
    "Jerman": "de",
    "Spanyol": "es",
    "Rusia": "ru",
    "Indonesia": "id"
}

# --- FUNGSI INTI TRANSLATE ---
def translate_core(text, target, source='id'):
    """Request ke Google API gtx."""
    if not text or str(text).strip().lower() in ["nan", "none", ""]:
        return ""
    
    headers = {"User-Agent": random.choice(USER_AGENTS)}
    base_url = "https://translate.googleapis.com/translate_a/single"
    
    params = {
        "client": "gtx",
        "sl": source,
        "tl": target,
        "dt": "t",
        "q": str(text).strip()
    }
    
    try:
        response = requests.get(base_url, params=params, headers=headers, timeout=12)
        if response.status_code == 200:
            result_json = response.json()
            # Google memecah hasil jika ada banyak kalimat, kita gabungkan kembali
            translated_parts = [part[0] for part in result_json[0] if part[0]]
            return "".join(translated_parts)
        elif response.status_code == 429:
            return "ERR_LIMIT"
    except Exception:
        pass
    return None

def translate_smart(text, target, use_indirect=False):
    """
    Logika Chunking & Indirect Translation.
    Jika use_indirect=True (untuk Dari), alur: ID -> EN -> Target.
    """
    text_str = str(text).strip()
    
    if not text_str:
        return ""

    # Tentukan alur terjemahan
    if use_indirect:
        # Langkah 1: Indonesia -> Inggris
        step1 = translate_core(text_str, 'en', source='id')
        if not step1 or step1 == "ERR_LIMIT":
            return step1 if step1 == "ERR_LIMIT" else None
        
        # Langkah 2: Inggris -> Target (fa-AF)
        # Kita pecah lagi jika hasil intermediate terlalu panjang, meski jarang terjadi
        if len(step1) <= 4500:
            final_result = translate_core(step1, target, source='en')
            return final_result
        else:
            # Jika hasil intermediate panjang, chunking di langkah 2
            chunks = [step1[i:i+4000] for i in range(0, len(step1), 4000)]
            translated_results = []
            for c in chunks:
                res = translate_core(c, target, source='en')
                if res and res != "ERR_LIMIT":
                    translated_results.append(res)
                else:
                    return "ERR_LIMIT" if res == "ERR_LIMIT" else None
            return " ".join(translated_results)

    else:
        # Langsung Translate (ID -> Target)
        if len(text_str) <= 4500:
            return translate_core(text_str, target, source='id')
        
        # Pecah per 4000 karakter agar aman dari limit URL
        chunks = [text_str[i:i+4000] for i in range(0, len(text_str), 4000)]
        translated_results = []
        
        for c in chunks:
            res = translate_core(c, target, source='id')
            if res and res != "ERR_LIMIT":
                translated_results.append(res)
            else:
                return "ERR_LIMIT" if res == "ERR_LIMIT" else None
                
        return " ".join(translated_results)

# --- ANTARMUKA STREAMLIT ---
st.set_page_config(page_title="Turbo Translator Pro v2", page_icon="⚡", layout="wide")

st.title("⚡ Turbo Excel Translator")
st.markdown("Alat translasi otomatis untuk file Excel buatan fadhil ganteng kece keren hebat slebew.  kalo gatau kodenya tanya gugel nulisnya gini 639-1 kode bahasa ..... bahasa mu ketiken. JANGAN LUPA DIKASIH LETI 1 BARIS DIATAS NYA")
st.markdown("PAKAILAH 1 TAB AJA JANGAN MULTI TAB WOYYYY RUSAK HOST E, NDAK TAK HOST NO MANEH WM")

# --- SIDEBAR ---
st.sidebar.header("⚙️ Pengaturan")

# Pilihan Bahasa yang Lebih User Friendly
lang_option = st.sidebar.selectbox(
    "Pilih Bahasa Tujuan",
    options=list(COMMON_LANGUAGES.keys()),
    index=1  # Default ke Dari (Afghanistan) karena konteks user sebelumnya
)

# Input manual jika ingin kode lain
use_custom_code = st.sidebar.checkbox("Gunakan Kode Bahasa Manual")
if use_custom_code:
    target_lang = st.sidebar.text_input("Masukkan Kode Bahasa (639-1)", value="en")
else:
    target_lang = COMMON_LANGUAGES[lang_option]

# Deteksi otomatis apakah perlu strategi "Indirect Translation"
# Strategi ID->EN->FA-AF biasanya lebih bagus untuk Dari karena data pelatihan EN->FA lebih banyak
is_dari_target = (target_lang == "fa-AF")
if is_dari_target:
    st.sidebar.info("🇦🇫 **Mode Dari Aktif:** Sistem akan menggunakan jalur terjemahan tidak langsung (ID→EN→FA-AF) untuk hasil yang lebih natural.")

max_workers = st.sidebar.slider("Kecepatan (Workers)", 1, 15, 5, help="Disarankan 5-10 agar aman.")

st.sidebar.markdown("---")
st.sidebar.info("📌 **Catatan:**\nJika hasil download berwarna merah, artinya IP kamu terkena limit sementara. Kurangi Workers atau ganti koneksi internet. pesan untuk mahrus UWES RUS NEK GA KUAT 10 AE GAUSA MEKSO DIULEK ULEK KODENE SAMPE DADI 100!!!")

# --- PROSES UTAMA ---
uploaded_file = st.file_uploader("Upload file Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)
        st.success(f"File dimuat: '{uploaded_file.name}' | Total: {len(df)} baris.")

        if st.button("🚀 Mulai Terjemahkan"):
            if df.shape[1] < 2:
                st.error("Kolom B (Kolom ke-2) tidak ditemukan!")
            else:
                texts_to_process = df.iloc[:, 1].tolist()
                total_rows = len(texts_to_process)
                results = [None] * total_rows
                
                # UI Progress
                progress_bar = st.progress(0)
                status_placeholder = st.empty()
                time_placeholder = st.empty()
                
                start_time = time.time()

                # --- MULTITHREADING ---
                # Kita kirim parameter use_indirect ke fungsi translate_smart
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    # Submit tasks dengan parameter tambahan use_indirect
                    future_to_idx = {
                        executor.submit(translate_smart, texts_to_process[i], target_lang, use_indirect=is_dari_target): i 
                        for i in range(total_rows)
                    }
                    
                    completed = 0
                    for future in future_to_idx:
                        idx = future_to_idx[future]
                        try:
                            results[idx] = future.result()
                        except:
                            results[idx] = None
                        
                        completed += 1
                        
                        # Hitung Estimasi Waktu
                        elapsed = time.time() - start_time
                        avg_time = elapsed / completed
                        eta = int(avg_time * (total_rows - completed))
                        
                        # Update Progress UI
                        progress_bar.progress(completed / total_rows)
                        status_placeholder.write(f"⏳ Memproses: {completed}/{total_rows} baris")
                        time_placeholder.markdown(f"⏱️ Sisa waktu: **{eta} detik**")

                df['Hasil Translate'] = results
                
                # --- PREVIEW ---
                st.subheader("📋 Preview Hasil (5 Baris Pertama)")
                st.dataframe(df[['Hasil Translate']].head(5))

                # --- GENERASI FILE DENGAN WARNA & NAMA DINAMIS ---
                # Logika Penamaan File: Nama_Asli (Kode_Bahasa).xlsx
                nama_file_murni = uploaded_file.name.rsplit('.', 1)[0]
                nama_file_baru = f"{nama_file_murni} ({target_lang}).xlsx"

                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Sheet1')
                    
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    # Cek error untuk diwarnai merah
                    for row_num, val in enumerate(results, start=2):
                        if val is None or val == "ERR_LIMIT" or val == "":
                            for col_num in range(1, df.shape[1] + 1):
                                worksheet.cell(row=row_num, column=col_num).fill = red_fill

                st.success(f"✅ Selesai! Nama file: {nama_file_baru}")
                
                st.download_button(
                    label="📥 Download Hasil Terjemahan",
                    data=output.getvalue(),
                    file_name=nama_file_baru,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:
        st.error(f"Terjadi kesalahan: {e}")
